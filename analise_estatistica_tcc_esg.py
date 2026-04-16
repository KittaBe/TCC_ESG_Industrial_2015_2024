# ==============================================================
# ANÁLISE ESTATÍSTICA - IMPACTO ESG NAS EMPRESAS INDUSTRIAIS
# ==============================================================
# - Lê a base Excel gerada pelo pipeline
# - Correlação, boxplots e regressões OLS com erros robustos (HC3)
# - Exporta: Tabelas no Excel (com interpretação), resumos .txt e log
# ==============================================================

import os
import numpy as np
import pandas as pd
import statsmodels.api as sm
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.styles import Font

# ------------------ CONFIG ------------------
BASE_DIR  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
BASE_PATH = os.path.join(BASE_DIR, "base_industrial_2015_2024.xlsx")
OUT_XLSX  = os.path.join(BASE_DIR, "resultados_estatisticos_2015_2024.xlsx")
OUT_TABELA = os.path.join(BASE_DIR, "tabela_resumida_anual_2015_2024.xlsx")
LOG_PATH  = os.path.join(BASE_DIR, "log_regressoes_2015_2024.txt")
RESUMO_TXT = os.path.join(BASE_DIR, "resumo_estatistico_2015_2024.txt")
FIG_DIR   = os.path.join(BASE_DIR, "outputs_2015_2024")
os.makedirs(FIG_DIR, exist_ok=True)

# ------------------ leitura ------------------
df = pd.read_excel(BASE_PATH, sheet_name="Base")
df.columns = df.columns.str.strip()

df = df.apply(pd.to_numeric, errors="ignore")
num_cols = df.select_dtypes(include=[np.number]).columns
df[num_cols] = df[num_cols].apply(pd.to_numeric, errors='coerce')

# ------------------ indicadores ------------------
df["ROE"] = df["Lucro Líquido"] / df["Patrimônio"]
df["ROA"] = df["Lucro Líquido"] / df["Ativos"]
df["Margem_Liquida"] = df["Lucro Líquido"] / df["Receita"]
df["Divida_PL"] = df["Passivo Total"] / df["Patrimônio"]
df["EV_EBITDA"] = df["EV (calc)"] / df["EBITDA (aprox)"]

# ------------------ estilo ------------------
sns.set_style("white")

def formatar(ax):
    ax.grid(False)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color('black')
    ax.spines['bottom'].set_color('black')
    ax.spines['left'].set_linewidth(1.5)
    ax.spines['bottom'].set_linewidth(1.5)

# ------------------ heatmap ------------------
corr = df[["Dummy ISE","ROE","ROA","Margem_Liquida","EV_EBITDA","Divida_PL"]].corr()

fig, ax = plt.subplots(figsize=(6,4))
sns.heatmap(corr, annot=True, fmt=".2f", cmap="coolwarm", vmin=-1, vmax=1, ax=ax)
formatar(ax)
plt.tight_layout()
plt.savefig(os.path.join(FIG_DIR, "figura1_heatmap.png"), dpi=300)
plt.close()

# ------------------ boxplots ------------------
for var, nome in [
    ("ROE","figura2_ROE.png"),
    ("Margem_Liquida","figura3_Margem.png"),
    ("EV_EBITDA","figura4_EV.png"),
    ("ROA","figura5_ROA.png"),
    ("Divida_PL","figura6_Divida.png")
]:
    fig, ax = plt.subplots(figsize=(6,4))
    sns.boxplot(x="Dummy ISE", y=var, data=df, ax=ax)
    formatar(ax)
    plt.tight_layout()
    plt.savefig(os.path.join(FIG_DIR, nome), dpi=300)
    plt.close()

# ------------------ regressões ------------------
def rodar(dep):
    X = df[["Dummy ISE","Ativos","Patrimônio"]].dropna()
    y = df[dep].loc[X.index]
    X = sm.add_constant(X)
    return sm.OLS(y, X).fit(cov_type="HC3")

modelos = {
    "ROE": rodar("ROE"),
    "ROA": rodar("ROA"),
    "Margem Líquida": rodar("Margem_Liquida"),
    "EV/EBITDA": rodar("EV_EBITDA"),
    "Dívida/PL": rodar("Divida_PL")
}

# ------------------ tabela principal ------------------
def estrelas(p):
    if p<0.01: return "***"
    elif p<0.05: return "**"
    elif p<0.10: return "*"
    else: return "ns"

def interpretar(nome, coef, p):
    direcao = "positivo" if coef > 0 else "negativo"
    if p < 0.05:
        return f"Efeito {direcao} e estatisticamente significativo do ESG sobre {nome}."
    else:
        return f"Efeito {direcao}, porém não significativo, do ESG sobre {nome}."

linhas = []
for nome, m in modelos.items():
    coef = m.params["Dummy ISE"]
    se   = m.bse["Dummy ISE"]
    p    = m.pvalues["Dummy ISE"]

    linhas.append({
        "Indicador": nome,
        "Coeficiente ESG": round(coef,3),
        "Erro padrão": round(se,3),
        "p-valor": round(p,3),
        "Significância": estrelas(p),
        "Interpretação": interpretar(nome, coef, p)
    })

tabela = pd.DataFrame(linhas)
tabela.to_excel(OUT_XLSX, index=False)

# ------------------ tabela resumo ------------------
resumo = df[[
    "Empresa","Ano","Receita","Lucro Líquido",
    "Ativos","Patrimônio","Passivo Total",
    "ROE","ROA","Margem_Liquida"
]].sort_values(["Empresa","Ano"])

resumo.to_excel(OUT_TABELA, index=False)

# ==============================================================
# EXPORTAÇÃO TXT (RESUMO ESTATÍSTICO)
# ==============================================================

with open(RESUMO_TXT, "w", encoding="utf-8") as f:
    f.write(">>> ANÁLISE ESG 2015–2024\n\n")

    f.write("Empresas na base:\n")
    f.write(str(df["Empresa"].value_counts()) + "\n\n")

    f.write("Estatísticas descritivas:\n")
    f.write(str(df[["ROE","ROA","Margem_Liquida","EV_EBITDA","Divida_PL"]].describe()) + "\n\n")

    f.write("Correlação:\n")
    f.write(str(corr) + "\n")

# ==============================================================
# LOG DAS REGRESSÕES
# ==============================================================

with open(LOG_PATH, "w", encoding="utf-8") as f:
    for nome, modelo in modelos.items():
        f.write(f"\n{'='*60}\n")
        f.write(f"MODELO: {nome}\n")
        f.write(f"{'='*60}\n\n")
        f.write(modelo.summary().as_text())
        f.write("\n\n")

# ------------------ formatação ------------------
wb = load_workbook(OUT_XLSX)
ws = wb.active

for col in range(1, ws.max_column+1):
    ws.cell(row=1, column=col).font = Font(bold=True)

wb.save(OUT_XLSX)

print("✅ ANÁLISE COMPLETA COM EXCEL + TXT + LOG")



