# ==============================================================
# PIPELINE TCC ESG (INDUSTRIAL) 2015–2024
# - Baixa DFP da CVM (DRE/BPA/BPP consolidados)
# - Busca dados de mercado (Yahoo Finance)
# - Calcula indicadores e cria Excel com FÓRMULAS e formatação
# ==============================================================

import os, zipfile, requests, time
import pandas as pd
import numpy as np
from datetime import date, timedelta
from tqdm import tqdm
from unidecode import unidecode
import yfinance as yf
from typing import Optional
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter

# -----------------------------
# CONFIGURAÇÕES
# -----------------------------
ANOS = list(range(2015, 2025))  # 2015–2024

BASE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
os.makedirs(BASE_DIR, exist_ok=True)

SAIDA_XLSX = os.path.join(BASE_DIR, "base_industrial_2015_2024.xlsx")
CACHE_DIR = os.path.join(BASE_DIR, "cvm_cache")

# Empresas-alvo
EMPRESAS = {
    "Klabin S.A.": {
        "aliases": ["KLABIN", "KLABIN S.A."],
        "ticker": "KLBN11.SA",
        "ise": 1
    },
    "Gerdau S.A.": {
        "aliases": ["GERDAU S.A.", "GERDAU"],
        "ticker": "GGBR4.SA",
        "ise": 1
    },
    "Metalúrgica Gerdau S.A.": {
        "aliases": ["METALURGICA GERDAU", "METALURGICA GERDAU S.A."],
        "ticker": "GOAU4.SA",
        "ise": 0
    },
    "Randon S.A. Implementos e Participações": {
        "aliases": ["RANDON", "RANDONCORP", "RANDON S.A.", "RANDON S.A. IMPLEMENTOS E PARTICIPACOES"],
        "ticker": "RAPT4.SA",
        "ise": 0
    },
    "Tupy S.A.": {
        "aliases": ["TUPY S.A.", "TUPY"],
        "ticker": "TUPY3.SA",
        "ise": 0
    },
    "WEG S.A.": {
        "aliases": ["WEG S.A.", "WEG"],
        "ticker": "WEGE3.SA",
        "ise": 1
    },
}

KW = {
    "receita":   [r"receita", r"vendas líquidas", r"receita operacional líquida", r"receita de venda"],
    "lucro":     [r"lucro.*(exerc|per[ií]odo)", r"preju[ií]zo.*(exerc|per[ií]odo)", r"lucro.*liquido"],
    "ativo":     [r"ativo total"],
    "pl":        [r"patrim[oô]nio l[ií]quido"],
    "passivo":   [r"passivo total"],
    "depamort":  [r"deprecia[cç][aã]o", r"amortiza[cç][aã]o"],
    "caixa":     [r"caixa", r"equivalentes de caixa", r"dispon[ií]vel"]
}

def norm(s: str) -> str:
    return unidecode(str(s or "")).upper().strip()

def match_empresa(denom: str, aliases: list[str]) -> bool:
    d = norm(denom)
    return any(a in d for a in [norm(x) for x in aliases])

def request_with_retry(url: str, retries: int = 3, timeout: int = 120) -> bytes:
    for i in range(retries):
        try:
            r = requests.get(url, timeout=timeout)
            r.raise_for_status()
            return r.content
        except Exception:
            if i == retries - 1:
                raise
            time.sleep(2 * (i + 1))
    return b""

def baixar_zip_dfp(ano: int) -> str:
    os.makedirs(CACHE_DIR, exist_ok=True)
    fname = f"dfp_cia_aberta_{ano}.zip"
    path = os.path.join(CACHE_DIR, fname)
    if not os.path.exists(path):
        url = f"https://dados.cvm.gov.br/dados/CIA_ABERTA/DOC/DFP/DADOS/{fname}"
        content = request_with_retry(url)
        with open(path, "wb") as f:
            f.write(content)
    return path

def carregar_csv(z: zipfile.ZipFile, marcador: str) -> Optional[pd.DataFrame]:
    for name in z.namelist():
        if marcador.lower() in name.lower() and name.lower().endswith(".csv"):
            with z.open(name) as f:
                return pd.read_csv(f, sep=";", encoding="latin1", low_memory=False)
    return None

def filtrar_empresa_ano(df: Optional[pd.DataFrame], aliases: list[str], ano: int) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    df = df.copy()
    if "DENOM_CIA" in df.columns:
        df = df[df["DENOM_CIA"].apply(lambda x: match_empresa(x, aliases))]
    for c in ["DT_REFER", "DT_FIM_EXERC"]:
        if c in df.columns:
            df["ANO"] = pd.to_datetime(df[c], errors="coerce").dt.year
            df = df[df["ANO"] == ano]
    return df

def pick_valor(df: pd.DataFrame, keywords: list[str]) -> Optional[float]:
    if df is None or df.empty or "DS_CONTA" not in df.columns or "VL_CONTA" not in df.columns:
        return None
    desc = df["DS_CONTA"].astype(str).apply(norm)
    mask = False
    for kw in keywords:
        mask = desc.str.contains(norm(kw), regex=True, na=False) | mask
    sel = df.loc[mask]
    if sel.empty:
        return None
    vals = pd.to_numeric(sel["VL_CONTA"], errors="coerce").replace([np.inf, -np.inf], np.nan).dropna()
    if vals.empty:
        return None
    return float(vals.loc[vals.abs().idxmax()])

def ultimo_dia_util(ano: int) -> date:
    d = date(ano, 12, 31)
    while d.weekday() >= 5:
        d -= timedelta(days=1)
    return d

def marketcap_historica(ticker: str, ano: int) -> Optional[float]:
    try:
        t = yf.Ticker(ticker)
        d = ultimo_dia_util(ano)
        hist = t.history(start=str(d - timedelta(days=10)), end=str(d + timedelta(days=5)))
        if not hist.empty:
            close = float(hist["Close"].iloc[-1])
            shares = t.fast_info.get("shares_outstanding") or t.info.get("sharesOutstanding")
            if shares and shares > 0:
                return close * float(shares)
    except Exception:
        pass
    return None

# -----------------------------
# PIPELINE
# -----------------------------
linhas = []

for ano in tqdm(ANOS, desc="Processando anos DFP/CVM"):
    zip_path = baixar_zip_dfp(ano)
    with zipfile.ZipFile(zip_path, "r") as z:
        dre = carregar_csv(z, "DRE_con")
        bpa = carregar_csv(z, "BPA_con")
        bpp = carregar_csv(z, "BPP_con")

        for empresa, meta in EMPRESAS.items():
            aliases, ticker, dummy_ise = meta["aliases"], meta["ticker"], meta["ise"]

            dre_y = filtrar_empresa_ano(dre, aliases, ano)
            bpa_y = filtrar_empresa_ano(bpa, aliases, ano)
            bpp_y = filtrar_empresa_ano(bpp, aliases, ano)

            receita = pick_valor(dre_y, KW["receita"])
            lucro   = pick_valor(dre_y, KW["lucro"])
            ativos  = pick_valor(bpa_y, KW["ativo"])
            pl      = pick_valor(bpp_y, KW["pl"])

            passivo_raw = pick_valor(bpp_y, KW["passivo"])

            # CORREÇÃO CONTÁBIL
            if ativos is not None and pl is not None:
                passivo = ativos - pl
            else:
                passivo = passivo_raw

            caixa   = pick_valor(bpa_y, KW["caixa"])
            depam   = pick_valor(dre_y, KW["depamort"])

            ebitda = (lucro or 0) + (depam or 0)
            mktcap = marketcap_historica(ticker, ano)
            div_liq = (passivo or 0) - (caixa or 0)
            ev = (mktcap or 0) + div_liq
            pe = (mktcap / lucro) if (mktcap and lucro and lucro != 0) else None

            linhas.append({
                "Empresa": empresa, "Ano": ano, "Receita": receita,
                "Lucro Líquido": lucro, "Ativos": ativos, "Patrimônio": pl,
                "Passivo Total": passivo, "Caixa e Equivalentes": caixa,
                "Depreciação/Amortização": depam, "EBITDA (aprox)": ebitda,
                "MarketCap (hist)": mktcap, "EV (calc)": ev, "P/E (calc)": pe,
                "Dummy ISE": int(dummy_ise)
            })

# -----------------------------
# EXPORTAR EXCEL (SEM ALTERAÇÃO)
# -----------------------------
df = pd.DataFrame(linhas).sort_values(["Empresa", "Ano"]).reset_index(drop=True)

wb = Workbook()
ws = wb.active
ws.title = "Base"

for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

col_map = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}

def COL(name):
    return get_column_letter(col_map[name])

headers_calc = ["Margem Líquida", "ROE", "ROA", "Dívida/Patrimônio", "EV/EBITDA"]
for h in headers_calc:
    ws.cell(row=1, column=ws.max_column+1, value=h)

first_calc_col = ws.max_column - len(headers_calc) + 1

for row in range(2, ws.max_row+1):
    ws.cell(row=row, column=first_calc_col + 0,
            value=f'=IF({COL("Receita")}{row}=0,"",{COL("Lucro Líquido")}{row}/{COL("Receita")}{row})')
    ws.cell(row=row, column=first_calc_col + 1,
            value=f'=IF({COL("Patrimônio")}{row}=0,"",{COL("Lucro Líquido")}{row}/{COL("Patrimônio")}{row})')
    ws.cell(row=row, column=first_calc_col + 2,
            value=f'=IF({COL("Ativos")}{row}=0,"",{COL("Lucro Líquido")}{row}/{COL("Ativos")}{row})')
    ws.cell(row=row, column=first_calc_col + 3,
            value=f'=IF({COL("Patrimônio")}{row}=0,"",{COL("Passivo Total")}{row}/{COL("Patrimônio")}{row})')
    ws.cell(row=row, column=first_calc_col + 4,
            value=f'=IF({COL("EBITDA (aprox)")}{row}=0,"",{COL("EV (calc)")}{row}/{COL("EBITDA (aprox)")}{row})')

for c in range(1, ws.max_column+1):
    ws.column_dimensions[get_column_letter(c)].width = 20

wb.save(SAIDA_XLSX)

print(f"✅ Excel gerado em: {SAIDA_XLSX}")
